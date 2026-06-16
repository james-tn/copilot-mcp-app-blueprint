"""Blueprint export generators: importable Blueprint JSON, PDF and Excel.

* ``build_blueprint_json`` — the "Download Blueprint" artifact: a structured,
  versioned JSON document of the whole application design (the analog of Pega's
  importable blueprint export).
* ``build_pdf`` — a human-readable PDF (pure-Python writer, no compiled deps).
* ``build_xlsx`` — a multi-sheet workbook (openpyxl).
"""

from __future__ import annotations

import io
import json
import re
from datetime import datetime, timezone
from typing import Any

from . import data, store


def safe_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_")
    return (s or "blueprint")[:80]


def _type_label(t: str) -> str:
    return data.STEP_TYPES.get(t, t)


# ── Download Blueprint (importable JSON artifact) ────────────────────────────

def build_blueprint_json(bp: dict[str, Any]) -> bytes:
    counts = store.blueprint_counts(bp)
    doc = {
        "$schema": "https://pega.example/blueprint/v1.json",
        "kind": "PegaBlueprint",
        "version": "1.0",
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "id": bp["id"],
        "title": bp["title"],
        "context": {
            "organization": bp["orgName"],
            "industry": bp["industry"],
            "subIndustry": bp["subIndustry"],
            "purpose": bp["purpose"],
            "description": bp["description"],
            "language": bp["language"],
            "location": bp["location"],
        },
        "workflows": bp["caseTypes"],
        "personas": bp["personas"],
        "dataModel": {
            "identity": bp["identity"],
            "inboundEvents": bp["inboundEvents"],
            "dataObjects": bp["dataObjects"],
            "integrations": bp["integrations"],
        },
        "architecture": counts,
    }
    return json.dumps(doc, indent=2).encode("utf-8")


# ── PDF (pure-Python, zero compiled dependencies) ───────────────────────────
#
# A minimal multi-page text PDF writer using the built-in Helvetica fonts. This
# avoids fpdf2/pillow (compiled) so the server builds cleanly on any host.

def _lat1(s: str) -> str:
    """PDF WinAnsi/latin-1 text; normalize common unicode punctuation."""
    repl = {"\u2014": "-", "\u2013": "-", "\u2019": "'", "\u2018": "'",
            "\u201c": '"', "\u201d": '"', "\u2022": "*", "\u2026": "...", "\u00b7": "-"}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def _pdf_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


# Approx Helvetica char width as a fraction of font size (conservative average).
_CHAR_W = 0.52


def _wrap(text: str, size: float, max_w: float) -> list[str]:
    max_chars = max(1, int(max_w / (size * _CHAR_W)))
    out: list[str] = []
    for raw in text.split("\n"):
        words = raw.split(" ")
        line = ""
        for w in words:
            cand = w if not line else f"{line} {w}"
            if len(cand) <= max_chars:
                line = cand
            else:
                if line:
                    out.append(line)
                while len(w) > max_chars:
                    out.append(w[:max_chars])
                    w = w[max_chars:]
                line = w
        out.append(line)
    return out


def build_pdf(bp: dict[str, Any]) -> bytes:
    # Page geometry (A4, points).
    PW, PH, MARGIN = 595.0, 842.0, 50.0
    usable_w = PW - 2 * MARGIN
    accent = "0.353 0.122 0.667"  # #5a1faa-ish Pega indigo/purple
    black = "0 0 0"
    grey = "0.43 0.43 0.43"

    items: list[tuple[str, str, float, str, float]] = []

    def add(text: str, font: str = "H", size: float = 11, color: str = black, gap: float = 0.0) -> None:
        for ln in _wrap(_lat1(text), size, usable_w):
            items.append((ln, font, size, color, gap))
            gap = 0.0  # gap only before the first wrapped line

    def h1(t: str) -> None:
        add(t, "B", 15, accent, gap=10)

    c = store.blueprint_counts(bp)
    add("Pega Blueprint", "B", 22, accent)
    add(bp["title"], "H", 13)
    add(f"{bp['subIndustry']} - {bp['industry']} - {bp['id']}", "H", 9, grey)

    h1("Application Context")
    add(f"Organization: {bp['orgName']}   |   Location: {bp['location']}   |   Language: {bp['language']}")
    add(f"Industry: {bp['industry']} / {bp['subIndustry']}   |   Purpose: {bp['purpose']}")
    add(bp["description"])
    add(
        f"Architecture: {c['caseTypes']} workflows, {c['stages']} stages, {c['steps']} steps "
        f"({c['automations']} automated), {c['dataObjects']} data objects, {c['personas']} personas."
    )

    h1("Workflows (Case Types)")
    for case in bp["caseTypes"]:
        tag = " [primary]" if case.get("primary") else ""
        add(f"{case['name']}{tag}", "B", 13, black, gap=8)
        add(case["description"], "H", 10)
        for stage in case["stages"]:
            add(f"  {stage['name']}  ({stage['kind']} stage)", "B", 11, accent, gap=4)
            for proc in stage.get("processes") or []:
                add(f"    Process: {proc['name']}", "B", 10, grey, gap=2)
                for s in proc["steps"]:
                    add(f"      - {s['name']}  [{_type_label(s['type'])}]", "H", 10)
            for s in stage.get("steps") or []:
                add(f"    - {s['name']}  [{_type_label(s['type'])}]", "H", 10)

    h1("Data & Integrations")
    add(f"User identity: {bp['identity']}")
    enabled = [e["name"] for e in bp["inboundEvents"] if e["enabled"]]
    add(f"Inbound events: {', '.join(enabled) or '-'}")
    add("Data objects: " + ", ".join(f"{o['name']} ({o['systemOfRecord']})" for o in bp["dataObjects"]))
    if bp["integrations"]:
        add("Integrations:", "B", 11, black, gap=4)
        for it in bp["integrations"]:
            add(f"  - {it['name']}: {it['purpose']}", "H", 10)

    h1("Personas")
    for p in bp["personas"]:
        add(p["name"], "B", 12, black, gap=4)
        add(p["description"], "H", 10)

    # Paginate into content streams.
    pages: list[str] = []
    cur: list[str] = []
    y = PH - MARGIN
    for text, font, size, color, gap in items:
        leading = size * 1.35
        y -= gap
        if y - leading < MARGIN:
            pages.append("\n".join(cur))
            cur = []
            y = PH - MARGIN
        y -= leading
        fref = "F2" if font == "B" else "F1"
        cur.append(
            f"{color} rg BT /{fref} {size:.1f} Tf 1 0 0 1 {MARGIN:.1f} {y:.1f} Tm "
            f"({_pdf_escape(text)}) Tj ET"
        )
    if cur:
        pages.append("\n".join(cur))
    if not pages:
        pages = [""]

    # Assemble the PDF objects.
    objs: list[bytes] = []

    def obj(b: str | bytes) -> int:
        objs.append(b.encode("latin-1") if isinstance(b, str) else b)
        return len(objs)  # 1-based object number

    font1 = obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>")
    font2 = obj("<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>")
    pages_obj_num = len(objs) + 1  # reserve next id for the Pages node
    objs.append(b"")  # placeholder for Pages

    kids: list[int] = []
    for content in pages:
        stream = content.encode("latin-1")
        content_num = obj(
            b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
        )
        page_num = obj(
            f"<< /Type /Page /Parent {pages_obj_num} 0 R /MediaBox [0 0 {PW:.0f} {PH:.0f}] "
            f"/Resources << /Font << /F1 {font1} 0 R /F2 {font2} 0 R >> >> "
            f"/Contents {content_num} 0 R >>"
        )
        kids.append(page_num)

    kids_refs = " ".join(f"{k} 0 R" for k in kids)
    objs[pages_obj_num - 1] = (
        f"<< /Type /Pages /Count {len(kids)} /Kids [{kids_refs}] >>".encode("latin-1")
    )
    catalog_num = obj(f"<< /Type /Catalog /Pages {pages_obj_num} 0 R >>")

    # Serialize with xref table.
    out = bytearray(b"%PDF-1.4\n")
    offsets = [0] * (len(objs) + 1)
    for i, body in enumerate(objs, start=1):
        offsets[i] = len(out)
        out += f"{i} 0 obj\n".encode("latin-1") + body + b"\nendobj\n"
    xref_pos = len(out)
    out += f"xref\n0 {len(objs) + 1}\n".encode("latin-1")
    out += b"0000000000 65535 f \n"
    for i in range(1, len(objs) + 1):
        out += f"{offsets[i]:010d} 00000 n \n".encode("latin-1")
    out += (
        f"trailer\n<< /Size {len(objs) + 1} /Root {catalog_num} 0 R >>\n"
        f"startxref\n{xref_pos}\n%%EOF"
    ).encode("latin-1")
    return bytes(out)


# ── Excel ────────────────────────────────────────────────────────────────────

def build_xlsx(bp: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)
    c = store.blueprint_counts(bp)

    ov = wb.active
    ov.title = "Overview"
    ov.append(["Field", "Value"])
    ov["A1"].font = bold
    ov["B1"].font = bold
    for k, v in [
        ("Title", bp["title"]), ("Blueprint ID", bp["id"]), ("Industry", bp["industry"]),
        ("Sub-industry", bp["subIndustry"]), ("Purpose", bp["purpose"]),
        ("Organization", bp["orgName"]), ("Location", bp["location"]), ("Language", bp["language"]),
        ("Description", bp["description"]), ("Identity", bp["identity"]),
        ("# Workflows", c["caseTypes"]), ("# Stages", c["stages"]), ("# Steps", c["steps"]),
        ("# Automations", c["automations"]), ("# Data objects", c["dataObjects"]),
        ("# Personas", c["personas"]), ("# Integrations", c["integrations"]),
    ]:
        ov.append([k, v])
    ov.column_dimensions["A"].width = 22
    ov.column_dimensions["B"].width = 95

    wf = wb.create_sheet("Workflows")
    wf.append(["Workflow", "Stage", "Stage kind", "Process", "Step", "Step type"])
    for cell in wf[1]:
        cell.font = bold
    for case in bp["caseTypes"]:
        for stage in case["stages"]:
            for proc in stage.get("processes") or []:
                for s in proc["steps"]:
                    wf.append([case["name"], stage["name"], stage["kind"], proc["name"],
                               s["name"], _type_label(s["type"])])
            for s in stage.get("steps") or []:
                wf.append([case["name"], stage["name"], stage["kind"], "",
                           s["name"], _type_label(s["type"])])
    for col, w in zip("ABCDEF", (26, 22, 12, 18, 30, 20)):
        wf.column_dimensions[col].width = w

    ps = wb.create_sheet("Personas")
    ps.append(["Name", "Description"])
    for cell in ps[1]:
        cell.font = bold
    for p in bp["personas"]:
        ps.append([p["name"], p["description"]])
    for col, w in zip("AB", (26, 110)):
        ps.column_dimensions[col].width = w

    dm = wb.create_sheet("Data & Integrations")
    dm.append(["Type", "Name", "Detail"])
    for cell in dm[1]:
        cell.font = bold
    dm.append(["Identity", bp["identity"], ""])
    for e in bp["inboundEvents"]:
        dm.append(["Inbound event", e["name"], "Enabled" if e["enabled"] else "Available"])
    for o in bp["dataObjects"]:
        dm.append(["Data object", o["name"], o["systemOfRecord"]])
    for it in bp["integrations"]:
        dm.append(["Integration", it["name"], it["purpose"]])
    for col, w in zip("ABC", (16, 30, 60)):
        dm.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
